"""
SQLite database helper for tracking picks history and performance metrics.
"""
import json
import sqlite3
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional, List, Dict
import time
import pandas as pd

DB_PATH = Path(__file__).parent / "picks_history.db"

# Cache for team schedule lookups (date_str -> set of team abbreviations that played)
_team_schedule_cache = {}
EXCEL_PATH = Path(__file__).parent / "nba_picks_tracker.xlsx"


def get_connection():
    """Get database connection with row factory."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    """Initialize the database schema."""
    conn = get_connection()
    cursor = conn.cursor()

    # Original picks table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS picks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            player TEXT NOT NULL,
            stat TEXT NOT NULL,
            line REAL NOT NULL,
            prediction REAL NOT NULL,
            direction TEXT NOT NULL,
            edge REAL NOT NULL,
            confidence REAL,
            opponent TEXT,
            is_home INTEGER,
            actual_result REAL,
            won INTEGER
        )
    """)

    # Add new columns if they don't exist (for migration)
    columns_to_add = [
        ("model_type", "TEXT DEFAULT 'unknown'"),
        ("game_date", "TEXT"),
        ("player_id", "INTEGER"),
        ("team_abbrev", "TEXT"),
        ("graded_at", "TEXT"),
        ("voided", "INTEGER DEFAULT 0"),  # 1 = DNP/voided, not counted in performance
        ("void_reason", "TEXT"),  # Reason for void (DNP, postponed, etc.)
        ("prob_over", "REAL"),  # ML probability of going OVER the line (0-100)
    ]

    # Get existing columns
    cursor.execute("PRAGMA table_info(picks)")
    existing_columns = {row[1] for row in cursor.fetchall()}

    for col_name, col_def in columns_to_add:
        if col_name not in existing_columns:
            cursor.execute(f"ALTER TABLE picks ADD COLUMN {col_name} {col_def}")

    # Game predictions table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS game_predictions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT NOT NULL,
            game_date TEXT NOT NULL,
            home_team TEXT NOT NULL,
            away_team TEXT NOT NULL,
            home_team_id INTEGER,
            away_team_id INTEGER,
            predicted_winner TEXT NOT NULL,
            home_win_prob REAL NOT NULL,
            away_win_prob REAL NOT NULL,
            confidence REAL,
            actual_winner TEXT,
            correct INTEGER,
            key_factors TEXT,
            model_version TEXT DEFAULT 'v1.0',
            graded_at TEXT
        )
    """)

    # Add extended_data column if missing (stores full prediction JSON for GET /today)
    cursor.execute("PRAGMA table_info(game_predictions)")
    gp_columns = {row[1] for row in cursor.fetchall()}
    if 'extended_data' not in gp_columns:
        cursor.execute("ALTER TABLE game_predictions ADD COLUMN extended_data TEXT")

    # Users table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id         TEXT PRIMARY KEY,
            email      TEXT UNIQUE NOT NULL,
            hashed_password TEXT NOT NULL,
            username   TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)

    # Add role to users if not present
    cursor.execute("PRAGMA table_info(users)")
    users_columns = {row[1] for row in cursor.fetchall()}
    if "role" not in users_columns:
        cursor.execute("ALTER TABLE users ADD COLUMN role TEXT NOT NULL DEFAULT 'user'")
    if "avatar_url" not in users_columns:
        cursor.execute("ALTER TABLE users ADD COLUMN avatar_url TEXT")

    # Add user_id to picks if not present
    cursor.execute("PRAGMA table_info(picks)")
    picks_columns = {row[1] for row in cursor.fetchall()}
    if "user_id" not in picks_columns:
        cursor.execute("ALTER TABLE picks ADD COLUMN user_id TEXT")

    conn.commit()
    conn.close()


# ── User functions ────────────────────────────────────────────

def create_user(user_id: str, email: str, hashed_password: str, username: str) -> dict:
    """Insert a new user row. Returns the created user dict."""
    conn = get_connection()
    cursor = conn.cursor()
    now = datetime.utcnow().isoformat()
    cursor.execute(
        "INSERT INTO users (id, email, hashed_password, username, created_at) VALUES (?, ?, ?, ?, ?)",
        (user_id, email, hashed_password, username, now)
    )
    conn.commit()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row)


def get_user_by_email(email: str) -> Optional[dict]:
    """Return user dict for given email, or None if not found."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE email = ?", (email,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def get_user_by_id(user_id: str) -> Optional[dict]:
    """Return user dict for given id, or None if not found."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def update_user_avatar(user_id: str, avatar_url: str) -> Optional[dict]:
    """Set avatar_url for user. Returns updated user dict."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET avatar_url = ? WHERE id = ?",
        (avatar_url, user_id)
    )
    conn.commit()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def clear_user_avatar(user_id: str) -> Optional[dict]:
    """Set avatar_url to NULL for user. Returns updated user dict."""
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute(
        "UPDATE users SET avatar_url = NULL WHERE id = ?",
        (user_id,)
    )
    conn.commit()
    cursor.execute("SELECT * FROM users WHERE id = ?", (user_id,))
    row = cursor.fetchone()
    conn.close()
    return dict(row) if row else None


def save_game_prediction(prediction_data: dict) -> int:
    """Save a game prediction to the database."""
    conn = get_connection()
    cursor = conn.cursor()

    key_factors = prediction_data.get('key_factors')
    if isinstance(key_factors, list):
        key_factors = json.dumps(key_factors)

    # Store full prediction payload for fast retrieval by GET /today
    extended_data = prediction_data.get('extended_data')
    if extended_data is None and prediction_data.get('matchup'):
        extended_data = json.dumps(prediction_data)

    cursor.execute("""
        INSERT INTO game_predictions (
            timestamp, game_date, home_team, away_team,
            home_team_id, away_team_id, predicted_winner,
            home_win_prob, away_win_prob, confidence,
            key_factors, model_version, extended_data
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        datetime.now().isoformat(),
        prediction_data.get('game_date'),
        prediction_data.get('home_team'),
        prediction_data.get('away_team'),
        prediction_data.get('home_team_id'),
        prediction_data.get('away_team_id'),
        prediction_data.get('predicted_winner'),
        prediction_data.get('home_win_prob'),
        prediction_data.get('away_win_prob'),
        prediction_data.get('confidence'),
        key_factors,
        prediction_data.get('model_version', 'v1.0'),
        extended_data,
    ))

    pred_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return pred_id


def get_todays_stored_predictions() -> list:
    """Get today's stored game predictions from the DB (no NBA API call)."""
    conn = get_connection()
    cursor = conn.cursor()

    today = datetime.now().strftime('%Y-%m-%d')
    # Fetch most recent prediction per matchup for today
    cursor.execute("""
        SELECT * FROM game_predictions
        WHERE game_date = ?
        ORDER BY timestamp DESC
    """, (today,))

    rows = cursor.fetchall()
    conn.close()

    seen = set()
    results = []
    for row in rows:
        d = dict(row)
        key = (d['home_team'], d['away_team'])
        if key in seen:
            continue
        seen.add(key)

        # Prefer full extended_data if available
        if d.get('extended_data'):
            try:
                pred = json.loads(d['extended_data'])
                pred['prediction_id'] = d['id']
                results.append(pred)
                continue
            except (json.JSONDecodeError, TypeError):
                pass

        # Fallback: reconstruct minimal structure
        key_factors = []
        if d.get('key_factors'):
            try:
                key_factors = json.loads(d['key_factors'])
            except (json.JSONDecodeError, TypeError):
                pass

        results.append({
            'matchup': {
                'home_team': {
                    'team_id': d.get('home_team_id', 0),
                    'team_abbrev': d['home_team'],
                    'team_name': d['home_team'],
                    'record': '?-?',
                    'off_rating': 0.0,
                    'def_rating': 0.0,
                    'net_rating': 0.0,
                    'pace': 0.0,
                },
                'away_team': {
                    'team_id': d.get('away_team_id', 0),
                    'team_abbrev': d['away_team'],
                    'team_name': d['away_team'],
                    'record': '?-?',
                    'off_rating': 0.0,
                    'def_rating': 0.0,
                    'net_rating': 0.0,
                    'pace': 0.0,
                },
                'game_date': d['game_date'],
                'game_time': None,
            },
            'predicted_winner': d['predicted_winner'],
            'home_win_prob': d['home_win_prob'],
            'away_win_prob': d['away_win_prob'],
            'confidence': d.get('confidence', 50.0),
            'key_factors': key_factors,
            'prediction_id': d['id'],
        })

    return results


def get_game_predictions(days: int = 7) -> list:
    """Get game predictions for the last N days."""
    conn = get_connection()
    cursor = conn.cursor()

    cutoff = (datetime.now() - timedelta(days=days)).isoformat()

    cursor.execute("""
        SELECT * FROM game_predictions
        WHERE timestamp >= ?
        ORDER BY timestamp DESC
    """, (cutoff,))

    rows = cursor.fetchall()
    conn.close()

    results = []
    for row in rows:
        d = dict(row)
        if d.get('key_factors'):
            try:
                d['key_factors'] = json.loads(d['key_factors'])
            except (json.JSONDecodeError, TypeError):
                d['key_factors'] = []
        else:
            d['key_factors'] = []
        results.append(d)

    return results


def get_pending_game_predictions() -> list:
    """Get game predictions that haven't been graded yet."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM game_predictions
        WHERE correct IS NULL
        ORDER BY game_date DESC
    """)

    rows = cursor.fetchall()
    conn.close()

    results = []
    for row in rows:
        d = dict(row)
        if d.get('key_factors'):
            try:
                d['key_factors'] = json.loads(d['key_factors'])
            except (json.JSONDecodeError, TypeError):
                d['key_factors'] = []
        else:
            d['key_factors'] = []
        results.append(d)

    return results


def grade_game_prediction(prediction_id: int, actual_winner: str):
    """Grade a game prediction with the actual winner."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT predicted_winner FROM game_predictions WHERE id = ?", (prediction_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return

    correct = 1 if row['predicted_winner'] == actual_winner else 0

    cursor.execute("""
        UPDATE game_predictions
        SET actual_winner = ?, correct = ?, graded_at = ?
        WHERE id = ?
    """, (actual_winner, correct, datetime.now().isoformat(), prediction_id))

    conn.commit()
    conn.close()


def auto_grade_game_predictions() -> dict:
    """Auto-grade pending game predictions using scoreboard data."""
    pending = get_pending_game_predictions()
    if not pending:
        return {'graded_count': 0, 'errors': [], 'results': []}

    from nba_api.stats.endpoints import scoreboardv2 as sb
    import time as _time

    graded_count = 0
    errors = []
    results = []
    today = datetime.now().date()

    for pred in pending:
        game_date = pred.get('game_date', '')
        if not game_date:
            continue

        try:
            pred_date = datetime.strptime(game_date[:10], '%Y-%m-%d').date()
            if pred_date >= today:
                continue  # Game hasn't happened yet

            # Fetch scoreboard for that date
            date_str = pred_date.strftime('%m/%d/%Y')
            scoreboard = sb.ScoreboardV2(game_date=date_str)
            _time.sleep(0.5)
            games_header = scoreboard.get_data_frames()[0]
            line_score = scoreboard.get_data_frames()[1]

            if games_header.empty:
                continue

            # Find the matching game
            home_team = pred['home_team']
            away_team = pred['away_team']

            for _, game in games_header.iterrows():
                status = game.get('GAME_STATUS_ID', 1)
                if status != 3:  # Not final
                    continue

                # Match by team abbreviations from line score
                game_id = game['GAME_ID']
                game_lines = line_score[line_score['GAME_ID'] == game_id]

                if game_lines.empty:
                    continue

                teams_in_game = set(game_lines['TEAM_ABBREVIATION'].values)
                if home_team in teams_in_game and away_team in teams_in_game:
                    # Found the game — determine winner
                    home_line = game_lines[game_lines['TEAM_ABBREVIATION'] == home_team]
                    away_line = game_lines[game_lines['TEAM_ABBREVIATION'] == away_team]

                    if home_line.empty or away_line.empty:
                        continue

                    home_pts = home_line.iloc[0].get('PTS', 0)
                    away_pts = away_line.iloc[0].get('PTS', 0)

                    if home_pts is None or away_pts is None:
                        continue

                    actual_winner = home_team if home_pts > away_pts else away_team

                    grade_game_prediction(pred['id'], actual_winner)
                    correct = pred['predicted_winner'] == actual_winner

                    results.append({
                        'home_team': home_team,
                        'away_team': away_team,
                        'predicted_winner': pred['predicted_winner'],
                        'actual_winner': actual_winner,
                        'correct': correct,
                        'home_pts': int(home_pts),
                        'away_pts': int(away_pts),
                    })
                    graded_count += 1
                    break

        except Exception as e:
            errors.append(f"Error grading {pred.get('home_team')} vs {pred.get('away_team')}: {str(e)}")

    return {
        'graded_count': graded_count,
        'errors': errors,
        'results': results,
    }


def get_game_accuracy_stats() -> dict:
    """Get accuracy statistics for game predictions."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM game_predictions WHERE correct IS NOT NULL")
    graded = [dict(row) for row in cursor.fetchall()]

    cursor.execute("SELECT COUNT(*) FROM game_predictions")
    total = cursor.fetchone()[0]

    conn.close()

    if not graded:
        return {
            'total_predictions': total,
            'graded_predictions': 0,
            'correct': 0,
            'incorrect': 0,
            'accuracy': 0.0,
            'by_confidence_range': {},
            'recent_streak': '0-0',
        }

    correct = sum(1 for p in graded if p['correct'] == 1)
    incorrect = len(graded) - correct
    accuracy = (correct / len(graded) * 100) if graded else 0.0

    # By confidence range
    by_conf = {}
    for low, high in [(50, 55), (55, 60), (60, 65), (65, 100)]:
        label = f"{low}-{high}%" if high < 100 else f"{low}%+"
        range_preds = [p for p in graded if low <= (p.get('confidence') or 50) < high]
        if range_preds:
            range_correct = sum(1 for p in range_preds if p['correct'] == 1)
            by_conf[label] = {
                'total': len(range_preds),
                'correct': range_correct,
                'accuracy': round(range_correct / len(range_preds) * 100, 1),
            }

    # Recent streak
    sorted_graded = sorted(graded, key=lambda p: p.get('timestamp', ''), reverse=True)
    recent_correct = 0
    recent_total = min(10, len(sorted_graded))
    for p in sorted_graded[:recent_total]:
        if p['correct'] == 1:
            recent_correct += 1

    return {
        'total_predictions': total,
        'graded_predictions': len(graded),
        'correct': correct,
        'incorrect': incorrect,
        'accuracy': round(accuracy, 1),
        'by_confidence_range': by_conf,
        'recent_streak': f"{recent_correct}-{recent_total - recent_correct} L{recent_total}",
    }


def save_pick(pick_data: dict) -> int:
    """
    Save a pick to the database.

    Args:
        pick_data: Dict with keys: player, stat, line, prediction, direction,
                   edge, confidence, opponent, is_home, model_type, game_date,
                   player_id, team_abbrev

    Returns:
        The ID of the inserted pick
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        INSERT INTO picks (timestamp, player, stat, line, prediction, direction,
                          edge, confidence, opponent, is_home, model_type,
                          game_date, player_id, team_abbrev, prob_over, user_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        datetime.now().isoformat(),
        pick_data.get('player'),
        pick_data.get('stat'),
        pick_data.get('line'),
        pick_data.get('prediction'),
        pick_data.get('direction'),
        pick_data.get('edge'),
        pick_data.get('confidence'),
        pick_data.get('opponent'),
        pick_data.get('is_home', 0),
        pick_data.get('model_type', 'unknown'),
        pick_data.get('game_date'),
        pick_data.get('player_id'),
        pick_data.get('team_abbrev'),
        pick_data.get('prob_over'),
        pick_data.get('user_id'),
    ))

    pick_id = cursor.lastrowid
    conn.commit()
    conn.close()

    return pick_id


def get_picks_history(days: int = 30, user_id: str = None) -> list:
    """
    Get picks history for the last N days.

    Args:
        days: Number of days to look back (default 30)
        user_id: If provided, only return picks for this user

    Returns:
        List of pick dicts
    """
    conn = get_connection()
    cursor = conn.cursor()

    cutoff = (datetime.now() - timedelta(days=days)).isoformat()

    if user_id:
        cursor.execute("""
            SELECT * FROM picks
            WHERE timestamp >= ? AND (voided IS NULL OR voided = 0) AND user_id = ?
            ORDER BY timestamp DESC
        """, (cutoff, user_id))
    else:
        cursor.execute("""
            SELECT * FROM picks
            WHERE timestamp >= ?
            ORDER BY timestamp DESC
        """, (cutoff,))

    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def get_all_picks() -> list:
    """Get all picks in the database."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM picks ORDER BY timestamp DESC")

    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def update_pick_result(pick_id: int, actual_result: float, line: float, direction: str):
    """
    Update a pick with the actual result and win/loss status.

    Args:
        pick_id: The pick ID to update
        actual_result: The actual stat value
        line: The betting line
        direction: "OVER" or "UNDER"
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Determine if the pick won
    if direction == "OVER":
        won = 1 if actual_result > line else 0
    else:
        won = 1 if actual_result < line else 0

    # Handle push (exactly hit the line)
    if actual_result == line:
        won = None  # Push

    cursor.execute("""
        UPDATE picks
        SET actual_result = ?, won = ?
        WHERE id = ?
    """, (actual_result, won, pick_id))

    conn.commit()
    conn.close()


def delete_pick(pick_id: int):
    """Delete a pick from the database."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM picks WHERE id = ?", (pick_id,))

    conn.commit()
    conn.close()


def void_pick(pick_id: int, reason: str = "DNP"):
    """
    Void a pick (DNP, postponed, etc.) - removes from performance calculations.

    Args:
        pick_id: The pick ID to void
        reason: Reason for voiding (DNP, postponed, injury, etc.)
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE picks
        SET voided = 1, void_reason = ?, won = NULL, actual_result = NULL
        WHERE id = ?
    """, (reason, pick_id))

    conn.commit()
    conn.close()


def get_voided_picks() -> List[Dict]:
    """Get all voided picks."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM picks
        WHERE voided = 1
        ORDER BY timestamp DESC
    """)

    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def unvoid_pick(pick_id: int):
    """Restore a voided pick back to pending status."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE picks
        SET voided = 0, void_reason = NULL
        WHERE id = ?
    """, (pick_id,))

    conn.commit()
    conn.close()


def reset_pick_to_pending(pick_id: int):
    """Reset a graded pick back to pending status (clears result)."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        UPDATE picks
        SET won = NULL, actual_result = NULL, graded_at = NULL
        WHERE id = ?
    """, (pick_id,))

    conn.commit()
    conn.close()


def reset_all_graded_for_date(game_date: str) -> int:
    """Reset all graded picks for a specific date back to pending.

    Args:
        game_date: Date string (YYYY-MM-DD format)

    Returns:
        Number of picks reset
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Count how many will be affected
    cursor.execute("""
        SELECT COUNT(*) FROM picks
        WHERE game_date LIKE ? AND won IS NOT NULL AND (voided IS NULL OR voided = 0)
    """, (f"{game_date}%",))
    count = cursor.fetchone()[0]

    # Reset them
    cursor.execute("""
        UPDATE picks
        SET won = NULL, actual_result = NULL, graded_at = NULL
        WHERE game_date LIKE ? AND won IS NOT NULL AND (voided IS NULL OR voided = 0)
    """, (f"{game_date}%",))

    conn.commit()
    conn.close()

    return count


def get_performance_stats(user_id: str = None) -> dict:
    """
    Calculate performance statistics (excludes voided picks).

    Returns:
        Dict with: total_picks, graded_picks, wins, losses, pushes,
                   win_rate, roi, avg_edge_winners, by_stat, by_edge_range
    """
    conn = get_connection()
    cursor = conn.cursor()

    # Get all graded picks (where won is not null, excludes voided)
    if user_id:
        cursor.execute("""
            SELECT * FROM picks WHERE won IS NOT NULL AND (voided IS NULL OR voided = 0) AND user_id = ?
        """, (user_id,))
    else:
        cursor.execute("""
            SELECT * FROM picks WHERE won IS NOT NULL AND (voided IS NULL OR voided = 0)
        """)
    graded = [dict(row) for row in cursor.fetchall()]

    # Get total non-voided picks
    if user_id:
        cursor.execute("SELECT COUNT(*) FROM picks WHERE (voided IS NULL OR voided = 0) AND user_id = ?", (user_id,))
    else:
        cursor.execute("SELECT COUNT(*) FROM picks WHERE voided IS NULL OR voided = 0")
    total_picks = cursor.fetchone()[0]

    conn.close()

    if not graded:
        return {
            'total_picks': total_picks,
            'graded_picks': 0,
            'wins': 0,
            'losses': 0,
            'pushes': 0,
            'win_rate': 0.0,
            'roi': 0.0,
            'avg_edge_winners': 0.0,
            'by_stat': {},
            'by_edge_range': {}
        }

    wins = sum(1 for p in graded if p['won'] == 1)
    losses = sum(1 for p in graded if p['won'] == 0)
    pushes = sum(1 for p in graded if p['won'] is None)

    # Win rate (excluding pushes)
    decided = wins + losses
    win_rate = (wins / decided * 100) if decided > 0 else 0.0

    # ROI calculation (assuming -110 odds, risk 1.1 to win 1.0)
    # Each win: +1.0 unit, each loss: -1.1 units
    profit = (wins * 1.0) - (losses * 1.1)
    total_risked = decided * 1.1
    roi = (profit / total_risked * 100) if total_risked > 0 else 0.0

    # Average edge on winners
    winners = [p for p in graded if p['won'] == 1]
    avg_edge_winners = sum(abs(p['edge']) for p in winners) / len(winners) if winners else 0.0

    # Performance by stat
    by_stat = {}
    for stat in ['PTS', 'REB', 'AST', 'PRA']:
        stat_picks = [p for p in graded if p['stat'] == stat and p['won'] is not None]
        if stat_picks:
            stat_wins = sum(1 for p in stat_picks if p['won'] == 1)
            stat_decided = sum(1 for p in stat_picks if p['won'] in [0, 1])
            by_stat[stat] = {
                'total': len(stat_picks),
                'wins': stat_wins,
                'win_rate': (stat_wins / stat_decided * 100) if stat_decided > 0 else 0.0
            }

    # Performance by edge range
    edge_ranges = [(5, 8), (8, 12), (12, 100)]
    by_edge_range = {}
    for low, high in edge_ranges:
        range_picks = [p for p in graded if low <= abs(p['edge']) < high and p['won'] is not None]
        if range_picks:
            range_wins = sum(1 for p in range_picks if p['won'] == 1)
            range_decided = sum(1 for p in range_picks if p['won'] in [0, 1])
            label = f"{low}-{high}%" if high < 100 else f"{low}%+"
            by_edge_range[label] = {
                'total': len(range_picks),
                'wins': range_wins,
                'win_rate': (range_wins / range_decided * 100) if range_decided > 0 else 0.0
            }

    return {
        'total_picks': total_picks,
        'graded_picks': len(graded),
        'wins': wins,
        'losses': losses,
        'pushes': pushes,
        'win_rate': win_rate,
        'roi': roi,
        'avg_edge_winners': avg_edge_winners,
        'by_stat': by_stat,
        'by_edge_range': by_edge_range
    }


def get_cumulative_profit(user_id: str = None) -> list:
    """
    Get cumulative profit over time for charting.

    Returns:
        List of dicts with: date, profit, cumulative_profit
    """
    conn = get_connection()
    cursor = conn.cursor()

    if user_id:
        cursor.execute("""
            SELECT timestamp, won FROM picks
            WHERE won IS NOT NULL AND user_id = ?
            ORDER BY timestamp ASC
        """, (user_id,))
    else:
        cursor.execute("""
            SELECT timestamp, won FROM picks
            WHERE won IS NOT NULL
            ORDER BY timestamp ASC
        """)

    rows = cursor.fetchall()
    conn.close()

    results = []
    cumulative = 0.0

    for row in rows:
        if row['won'] == 1:
            profit = 1.0
        elif row['won'] == 0:
            profit = -1.1
        else:
            profit = 0.0

        cumulative += profit
        results.append({
            'date': row['timestamp'][:10],  # Just the date part
            'profit': profit,
            'cumulative_profit': round(cumulative, 2)
        })

    return results


def get_pending_picks(user_id: str = None) -> List[Dict]:
    """Get all picks that haven't been graded yet (excludes voided picks)."""
    conn = get_connection()
    cursor = conn.cursor()

    if user_id:
        cursor.execute("""
            SELECT * FROM picks
            WHERE won IS NULL AND (voided IS NULL OR voided = 0) AND user_id = ?
            ORDER BY timestamp DESC
        """, (user_id,))
    else:
        cursor.execute("""
            SELECT * FROM picks
            WHERE won IS NULL AND (voided IS NULL OR voided = 0)
            ORDER BY timestamp DESC
        """)

    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def get_picks_for_date(game_date: str) -> List[Dict]:
    """Get all picks for a specific game date."""
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT * FROM picks
        WHERE game_date = ?
        ORDER BY timestamp DESC
    """, (game_date,))

    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def _check_team_played(team_abbrev: str, game_date, scraper=None) -> bool:
    """
    Check if a team played on a specific date using the NBA scoreboard API.

    Args:
        team_abbrev: Team abbreviation (e.g., 'LAL', 'BOS')
        game_date: date object for the game
        scraper: Optional NBADataScraper instance

    Returns:
        True if the team played on that date, False otherwise
    """
    if not team_abbrev:
        return False

    date_str = game_date.strftime('%m/%d/%Y')

    # Check cache first
    if date_str in _team_schedule_cache:
        return team_abbrev.upper() in _team_schedule_cache[date_str]

    try:
        from nba_api.stats.endpoints import scoreboardv2
        scoreboard = scoreboardv2.ScoreboardV2(game_date=date_str)
        games = scoreboard.get_data_frames()[0]  # GameHeader

        teams_that_played = set()
        if not games.empty:
            for _, game in games.iterrows():
                home = str(game.get('HOME_TEAM_ID', ''))
                away = str(game.get('VISITOR_TEAM_ID', ''))
                # Also grab abbreviations from the line score table
            # Use line score for abbreviations
            line_score = scoreboard.get_data_frames()[1]
            if not line_score.empty and 'TEAM_ABBREVIATION' in line_score.columns:
                for abbrev in line_score['TEAM_ABBREVIATION'].unique():
                    teams_that_played.add(str(abbrev).upper())

        _team_schedule_cache[date_str] = teams_that_played
        time.sleep(0.5)  # Rate limiting
        return team_abbrev.upper() in teams_that_played

    except Exception:
        return False


def auto_void_stale_picks(days_threshold: int = 3) -> int:
    """
    Auto-void picks that are stuck in pending for too long (safety net).

    Args:
        days_threshold: Number of days after which a pending pick is considered stale

    Returns:
        Number of picks voided
    """
    conn = get_connection()
    cursor = conn.cursor()

    cutoff_date = (datetime.now() - timedelta(days=days_threshold)).strftime('%Y-%m-%d')

    cursor.execute("""
        SELECT id, player, stat, game_date FROM picks
        WHERE won IS NULL AND (voided IS NULL OR voided = 0)
        AND game_date IS NOT NULL AND game_date < ?
    """, (cutoff_date,))

    stale_picks = cursor.fetchall()
    conn.close()

    voided = 0
    for pick in stale_picks:
        void_pick(pick['id'], "DNP")
        voided += 1

    return voided


def get_stale_pending_picks(days_threshold: int = 2) -> List[Dict]:
    """
    Get pending picks that are older than the given threshold.

    Args:
        days_threshold: Number of days to consider a pick stale

    Returns:
        List of stale pending picks
    """
    conn = get_connection()
    cursor = conn.cursor()

    cutoff_date = (datetime.now() - timedelta(days=days_threshold)).strftime('%Y-%m-%d')

    cursor.execute("""
        SELECT * FROM picks
        WHERE won IS NULL AND (voided IS NULL OR voided = 0)
        AND game_date IS NOT NULL AND game_date < ?
        ORDER BY game_date ASC
    """, (cutoff_date,))

    rows = cursor.fetchall()
    conn.close()
    return [dict(row) for row in rows]


def auto_grade_picks(scraper=None) -> Dict:
    """
    Automatically grade pending picks by fetching actual results from NBA API.

    Args:
        scraper: Optional NBADataScraper instance (will create one if not provided)

    Returns:
        Dict with: graded_count, errors, results
    """
    # Import here to avoid circular imports
    if scraper is None:
        from nba_evaluator import NBADataScraper
        scraper = NBADataScraper()

    # Safety net: auto-void picks that are 3+ days old
    stale_voided = auto_void_stale_picks(days_threshold=3)

    pending = get_pending_picks()
    if not pending:
        return {'graded_count': 0, 'errors': [], 'results': [], 'stale_voided': stale_voided}

    graded_count = 0
    errors = []
    results = []

    # Group picks by player to reduce API calls
    players_processed = {}

    # Get today's date for comparison
    today = datetime.now().date()

    for pick in pending:
        player_name = pick['player']
        player_id = pick.get('player_id')
        game_date = pick.get('game_date')
        stat = pick['stat']

        try:
            # Skip picks for future games (game date is after today)
            if game_date:
                try:
                    pick_date = datetime.strptime(game_date[:10], '%Y-%m-%d').date()
                    if pick_date > today:
                        # Game is in the future, skip
                        continue
                except ValueError:
                    pass

            # Get player ID if not stored
            if not player_id:
                player_info = scraper.get_player_info(player_name)
                if not player_info:
                    errors.append(f"Could not find player: {player_name}")
                    continue
                player_id = player_info['player_id']

            # Get game log (use cached if we already fetched for this player)
            # ONLY fetch current season to avoid matching old games for DNP players
            if player_id not in players_processed:
                game_log = scraper.get_player_game_log(player_id, seasons=['2025-26'])
                players_processed[player_id] = game_log
                time.sleep(0.5)  # Rate limiting
            else:
                game_log = players_processed[player_id]

            # Find the game for this pick
            opponent = pick.get('opponent', '')
            game_match = None
            player_dnp = False

            if game_date:
                try:
                    # Convert ISO date to NBA API format (e.g., "2026-02-03" -> "Feb 03, 2026")
                    dt = datetime.strptime(game_date[:10], '%Y-%m-%d')
                    nba_date_format = dt.strftime('%b %d, %Y')

                    # Check if player has any games this season
                    if game_log is None or game_log.empty:
                        # Player has no games at all this season
                        # Check if the game date is in the past - if so, it's a DNP
                        if dt.date() <= today:
                            player_dnp = True
                    else:
                        # Normalize game log dates to YYYY-MM-DD for reliable comparison
                        game_log['GAME_DATE_NORM'] = pd.to_datetime(game_log['GAME_DATE']).dt.strftime('%Y-%m-%d')
                        pick_date_str = dt.strftime('%Y-%m-%d')

                        # First try exact date match
                        game_match = game_log[game_log['GAME_DATE_NORM'] == pick_date_str]

                        # Filter by opponent if we have one and multiple games
                        if not game_match.empty and opponent and len(game_match) > 1:
                            opponent_match = game_match[game_match['MATCHUP'].str.contains(opponent)]
                            if not opponent_match.empty:
                                game_match = opponent_match

                        # If no exact date match, try +/- 1 day (games sometimes shift dates due to timezone)
                        # But REQUIRE opponent match to avoid grading wrong games
                        if game_match.empty and opponent:
                            dt_minus1 = (dt - timedelta(days=1)).strftime('%Y-%m-%d')
                            dt_plus1 = (dt + timedelta(days=1)).strftime('%Y-%m-%d')
                            nearby_games = game_log[game_log['GAME_DATE_NORM'].isin([dt_minus1, dt_plus1])]

                            # Only use nearby date if opponent also matches
                            if not nearby_games.empty:
                                opponent_match = nearby_games[nearby_games['MATCHUP'].str.contains(opponent)]
                                if not opponent_match.empty:
                                    game_match = opponent_match

                        # If no game found but date is in the past, check if team played
                        if (game_match is None or game_match.empty) and dt.date() <= today:
                            days_since_game = (today - dt.date()).days
                            if days_since_game >= 1:
                                # Check if the team actually played that day
                                team_played = _check_team_played(
                                    pick.get('team_abbrev'), dt.date(), scraper
                                )
                                if team_played:
                                    # Team played but player has no stats → confirmed DNP
                                    player_dnp = True
                                elif days_since_game >= 2:
                                    # Fallback: 2-day rule for cases where team check fails
                                    player_dnp = True
                            # else: game was today, leave as pending

                except ValueError:
                    # Date parsing failed - skip this pick
                    continue

            # Handle DNP - void the pick automatically
            if player_dnp:
                void_pick(pick['id'], "DNP")
                results.append({
                    'player': player_name,
                    'stat': stat,
                    'line': pick['line'],
                    'prediction': pick['prediction'],
                    'actual': None,
                    'direction': pick['direction'],
                    'won': None,
                    'voided': True,
                    'void_reason': 'DNP',
                    'model_type': pick.get('model_type', 'unknown')
                })
                continue

            # NO FALLBACK - if we can't find a date match, the game hasn't happened yet
            # Do NOT match by opponent only as this would grade using old games
            if game_match is None or game_match.empty:
                # Game not found for this date - likely hasn't been played yet
                continue

            # Get the actual stat value
            game = game_match.iloc[0]

            if stat == 'PRA':
                actual = game['PTS'] + game['REB'] + game['AST']
            elif stat in game:
                actual = game[stat]
            else:
                errors.append(f"Stat {stat} not found for {player_name}")
                continue

            # Update the pick
            update_pick_result(pick['id'], float(actual), pick['line'], pick['direction'])

            # Mark as graded
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE picks SET graded_at = ? WHERE id = ?",
                          (datetime.now().isoformat(), pick['id']))
            conn.commit()
            conn.close()

            # Determine result
            line = pick['line']
            direction = pick['direction']
            if direction == "OVER":
                won = actual > line
            else:
                won = actual < line

            results.append({
                'player': player_name,
                'stat': stat,
                'line': line,
                'prediction': pick['prediction'],
                'actual': actual,
                'direction': direction,
                'won': won,
                'model_type': pick.get('model_type', 'unknown')
            })

            graded_count += 1

        except Exception as e:
            errors.append(f"Error grading {player_name}: {str(e)}")

    # Count voided picks
    voided_count = sum(1 for r in results if r.get('voided'))

    return {
        'graded_count': graded_count,
        'voided_count': voided_count,
        'stale_voided': stale_voided,
        'errors': errors,
        'results': results
    }


def get_performance_by_model() -> Dict:
    """
    Get performance statistics broken down by model type.

    Returns:
        Dict with model types as keys, each containing win_rate, total, wins, roi
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT model_type, won, edge FROM picks
        WHERE won IS NOT NULL AND (voided IS NULL OR voided = 0)
    """)

    rows = cursor.fetchall()
    conn.close()

    # Group by model type
    model_stats = {}
    for row in rows:
        model = row['model_type'] or 'unknown'
        if model not in model_stats:
            model_stats[model] = {'wins': 0, 'losses': 0, 'total': 0, 'edges': []}

        model_stats[model]['total'] += 1
        if row['won'] == 1:
            model_stats[model]['wins'] += 1
            model_stats[model]['edges'].append(abs(row['edge']))
        elif row['won'] == 0:
            model_stats[model]['losses'] += 1

    # Calculate win rates and ROI for each model
    results = {}
    for model, stats in model_stats.items():
        decided = stats['wins'] + stats['losses']
        win_rate = (stats['wins'] / decided * 100) if decided > 0 else 0.0

        # ROI calculation (assuming -110 odds)
        profit = (stats['wins'] * 1.0) - (stats['losses'] * 1.1)
        total_risked = decided * 1.1
        roi = (profit / total_risked * 100) if total_risked > 0 else 0.0

        avg_edge = sum(stats['edges']) / len(stats['edges']) if stats['edges'] else 0.0

        results[model] = {
            'total': stats['total'],
            'wins': stats['wins'],
            'losses': stats['losses'],
            'win_rate': round(win_rate, 1),
            'roi': round(roi, 1),
            'avg_edge_winners': round(avg_edge, 1)
        }

    return results


def get_performance_by_model_and_stat() -> Dict:
    """
    Get detailed performance breakdown by model type AND stat type.

    Returns:
        Nested dict: {model_type: {stat: {win_rate, total, wins}}}
    """
    conn = get_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT model_type, stat, won, edge FROM picks
        WHERE won IS NOT NULL AND (voided IS NULL OR voided = 0)
    """)

    rows = cursor.fetchall()
    conn.close()

    # Group by model and stat
    data = {}
    for row in rows:
        model = row['model_type'] or 'unknown'
        stat = row['stat']

        if model not in data:
            data[model] = {}
        if stat not in data[model]:
            data[model][stat] = {'wins': 0, 'losses': 0, 'total': 0}

        data[model][stat]['total'] += 1
        if row['won'] == 1:
            data[model][stat]['wins'] += 1
        elif row['won'] == 0:
            data[model][stat]['losses'] += 1

    # Calculate win rates
    results = {}
    for model, stats in data.items():
        results[model] = {}
        for stat, counts in stats.items():
            decided = counts['wins'] + counts['losses']
            win_rate = (counts['wins'] / decided * 100) if decided > 0 else 0.0
            results[model][stat] = {
                'total': counts['total'],
                'wins': counts['wins'],
                'losses': counts['losses'],
                'win_rate': round(win_rate, 1)
            }

    return results


def export_to_excel() -> str:
    """
    Export all picks and performance stats to an Excel file with multiple sheets.

    Returns:
        Path to the created Excel file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
    except ImportError:
        print("⚠️ openpyxl not installed. Run: pip install openpyxl")
        return None

    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ==================== SHEET 1: All Picks ====================
    ws_picks = wb.active
    ws_picks.title = "All Picks"

    all_picks = get_all_picks()

    # Headers
    headers = ["ID", "Date", "Player", "Stat", "Direction", "Line", "Prediction", "Edge %",
               "Opponent", "Home/Away", "Model", "Result", "Actual", "Won"]
    for col, header in enumerate(headers, 1):
        cell = ws_picks.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, pick in enumerate(all_picks, 2):
        data = [
            pick['id'],
            pick['timestamp'][:10] if pick.get('timestamp') else '',
            pick['player'],
            pick['stat'],
            pick['direction'],
            pick['line'],
            round(pick['prediction'], 1),
            round(pick['edge'], 1) if pick.get('edge') else 0,
            pick.get('opponent', ''),
            'HOME' if pick.get('is_home') else 'AWAY',
            (pick.get('model_type') or 'unknown').replace('_', ' ').title(),
            'WIN' if pick.get('won') == 1 else 'LOSS' if pick.get('won') == 0 else 'PENDING',
            round(pick['actual_result'], 1) if pick.get('actual_result') is not None else '',
            pick.get('won', '')
        ]

        for col, value in enumerate(data, 1):
            cell = ws_picks.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Color code by result
            if pick.get('won') == 1:
                cell.fill = win_fill
            elif pick.get('won') == 0:
                cell.fill = loss_fill
            elif pick.get('won') is None:
                cell.fill = pending_fill

    # Auto-adjust column widths
    for col in ws_picks.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_picks.column_dimensions[column].width = min(max_length + 2, 25)

    # ==================== SHEET 2: Pending Picks ====================
    ws_pending = wb.create_sheet("Pending Picks")

    pending = get_pending_picks()
    pending_sorted = sorted(pending, key=lambda x: abs(x.get('edge', 0)), reverse=True)

    headers_pending = ["Player", "Stat", "Direction", "Line", "Prediction", "Edge %",
                       "Opponent", "Game Date", "Model"]
    for col, header in enumerate(headers_pending, 1):
        cell = ws_pending.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    for row_idx, pick in enumerate(pending_sorted, 2):
        data = [
            pick['player'],
            pick['stat'],
            pick['direction'],
            pick['line'],
            round(pick['prediction'], 1),
            round(pick['edge'], 1) if pick.get('edge') else 0,
            pick.get('opponent', ''),
            pick.get('game_date', '')[:10] if pick.get('game_date') else '',
            (pick.get('model_type') or 'unknown').replace('_', ' ').title()
        ]

        for col, value in enumerate(data, 1):
            cell = ws_pending.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Highlight high edge picks
            edge = pick.get('edge', 0)
            if abs(edge) >= 30:
                cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif abs(edge) >= 20:
                cell.fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")

    for col in ws_pending.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws_pending.column_dimensions[column].width = min(max_length + 2, 25)

    # ==================== SHEET 3: Performance Summary ====================
    ws_perf = wb.create_sheet("Performance")

    stats = get_performance_stats()
    model_stats = get_performance_by_model()

    # Overall stats
    ws_perf.cell(row=1, column=1, value="OVERALL PERFORMANCE").font = Font(bold=True, size=14)
    ws_perf.merge_cells('A1:C1')

    overall_data = [
        ("Total Picks", stats['total_picks']),
        ("Graded Picks", stats['graded_picks']),
        ("Wins", stats['wins']),
        ("Losses", stats['losses']),
        ("Win Rate", f"{stats['win_rate']:.1f}%"),
        ("ROI", f"{stats['roi']:+.1f}%"),
        ("Avg Edge (Winners)", f"{stats['avg_edge_winners']:.1f}%"),
        ("Last Updated", datetime.now().strftime("%Y-%m-%d %H:%M"))
    ]

    for row_idx, (label, value) in enumerate(overall_data, 3):
        ws_perf.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws_perf.cell(row=row_idx, column=2, value=value)

    # By Model
    ws_perf.cell(row=13, column=1, value="PERFORMANCE BY MODEL").font = Font(bold=True, size=14)
    ws_perf.merge_cells('A13:E13')

    model_headers = ["Model", "Wins", "Losses", "Win Rate", "ROI"]
    for col, header in enumerate(model_headers, 1):
        cell = ws_perf.cell(row=15, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 16
    for model, data in sorted(model_stats.items(), key=lambda x: x[1]['win_rate'], reverse=True):
        model_name = model.replace('_', ' ').title()
        row_data = [model_name, data['wins'], data['losses'], f"{data['win_rate']:.1f}%", f"{data['roi']:+.1f}%"]
        for col, value in enumerate(row_data, 1):
            cell = ws_perf.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
        row_idx += 1

    # By Stat
    ws_perf.cell(row=row_idx + 2, column=1, value="PERFORMANCE BY STAT").font = Font(bold=True, size=14)
    stat_row = row_idx + 4

    stat_headers = ["Stat", "Wins", "Total", "Win Rate"]
    for col, header in enumerate(stat_headers, 1):
        cell = ws_perf.cell(row=stat_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    stat_row += 1
    for stat, data in stats.get('by_stat', {}).items():
        row_data = [stat, data['wins'], data['total'], f"{data['win_rate']:.1f}%"]
        for col, value in enumerate(row_data, 1):
            cell = ws_perf.cell(row=stat_row, column=col, value=value)
            cell.border = thin_border
        stat_row += 1

    # By Edge Range
    ws_perf.cell(row=stat_row + 2, column=1, value="PERFORMANCE BY EDGE RANGE").font = Font(bold=True, size=14)
    edge_row = stat_row + 4

    edge_headers = ["Edge Range", "Wins", "Total", "Win Rate"]
    for col, header in enumerate(edge_headers, 1):
        cell = ws_perf.cell(row=edge_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    edge_row += 1
    for edge_range, data in stats.get('by_edge_range', {}).items():
        row_data = [edge_range, data['wins'], data['total'], f"{data['win_rate']:.1f}%"]
        for col, value in enumerate(row_data, 1):
            cell = ws_perf.cell(row=edge_row, column=col, value=value)
            cell.border = thin_border
        edge_row += 1

    ws_perf.column_dimensions['A'].width = 20
    ws_perf.column_dimensions['B'].width = 12
    ws_perf.column_dimensions['C'].width = 12
    ws_perf.column_dimensions['D'].width = 12
    ws_perf.column_dimensions['E'].width = 12

    # ==================== SHEET 4: Profit Tracker ====================
    ws_profit = wb.create_sheet("Profit Tracker")

    profit_data = get_cumulative_profit()

    headers_profit = ["Date", "Profit/Loss", "Cumulative Profit"]
    for col, header in enumerate(headers_profit, 1):
        cell = ws_profit.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, entry in enumerate(profit_data, 2):
        data = [entry['date'], entry['profit'], entry['cumulative_profit']]
        for col, value in enumerate(data, 1):
            cell = ws_profit.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col == 2:  # Profit column
                if value > 0:
                    cell.fill = win_fill
                elif value < 0:
                    cell.fill = loss_fill

    for col in ws_profit.columns:
        ws_profit.column_dimensions[col[0].column_letter].width = 18

    # Save the workbook
    wb.save(EXCEL_PATH)

    return str(EXCEL_PATH)


# Initialize database on import
init_db()
